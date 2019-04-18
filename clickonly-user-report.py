#**********************
# author: Nolan B. Kennedy (nxkennedy)
# THIS IS A JANKY AD HOC REPORTING SCRIPT. USE MY OTHER ONE FOR THE REAL GOODNESS.
# THIS ONE IS FOR REPORTING USERS THAT CLICKED DURING CLICK-ONLY SCENARIOS
# USAGE: python3 clickonly-user-report.py <database> <component_name>
# EXAMPLE: python3 clickonly-user-report.py db-20181012-140940.db CRT
#**********************

from openpyxl import Workbook
import sqlite3
import sys


db = sys.argv[1]
name = sys.argv[2]
name = name.upper()
dest_xlsx = name + '.xlsx'
conn = sqlite3.connect(db)
c = conn.cursor()
# Now let's get our workbook set up with all our sheets
wb = Workbook()

# Workbook saving
def save():
        wb.save(filename = dest_xlsx)


def writeToSheet(header, sheet, results):

	for col in range(0, len(header)):
		sheet.cell(row=1, column=col + 1).value = header[col]
	for row in results:
		sheet.append(row)
	save()


def rawData(name):

	print("[+] Writing data to '{}'...".format(dest_xlsx))
	wb.active.title = "Raw Data"
	sheet = wb.active
	header = ['Email', 'Recipient Name', 'Recipient Group', 'Department', 'Location', 'Opened Email Timestamp', 'Clicked Link?']
	query = '''
        SELECT Email, "Recipient Name", "Recipient Group", Department, Location, "Opened Email Timestamp", "Clicked Link?"
	FROM Reports WHERE("component" = "{}" AND "Clicked Link?" = "Yes" AND scenario_type = "Click-only")
	ORDER BY Email
	'''.format(name)
	results = c.execute(query)
	writeToSheet(header, sheet, results)


def clickStats(name):

	header = ['Email', 'Recipient Name', 'Department', 'Occurrences']
	title = 'Number of Occurrences'
	sheet = wb.create_sheet(title=title)
	query = '''
	SELECT Email, "Recipient Name", Department, COUNT(Email) as NumberOfOccurrence 
	FROM Reports WHERE("component" = "{}" AND "Clicked Link?" = "Yes" AND scenario_type = "Click-only") 
	GROUP BY Email
	ORDER BY NumberOfOccurrence DESC
        '''.format(name)
	results = c.execute(query)
	writeToSheet(header, sheet, results)


if __name__ == '__main__':

	rawData(name)
	clickStats(name)
	conn.close()
	print('[+] DONE. Spreadsheet Created.')
