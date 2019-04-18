# author: Nolan B. Kennedy (nxkennedy)

import csv
import sqlite3
from openpyxl import Workbook
import sys
import time



dest_xlsx = 'charts.xlsx'
export = 'db-export-' + time.strftime("%Y%m%d-%H%M%S") + '.csv'
components = []
db = sys.argv[1]
conn = sqlite3.connect(db)
c = conn.cursor()
# Now let's get our workbook set up with all our sheets
wb = Workbook()


# KNOWN GOOD. USE AS TEMPLATE.(THE 'ORDER BY DATE' FUNCTION IS OFF THOUGH DUE TO THE NON-DATE DATA FORMAT)
sql_response_rates_per_scenario_all_components = '''
	with total as
		(SELECT scenario_id as id, COUNT("Email") as total
		FROM Reports
		WHERE ("Last Email Status" NOT LIKE "%bounced%")
		GROUP BY id
		)

	SELECT scenario_id, component, scenario_type, "Last Email Status Timestamp" as date, total.total as total, COUNT("Email") as response, ROUND(CAST(COUNT("Email") AS FLOAT) / total.total, 3) as response_rate
	FROM Reports, total
	WHERE (total.id = scenario_id AND "Last Email Status" NOT LIKE "%bounced%")
        AND (("scenario_type" = "Click-only" AND "Clicked Link?" = "Yes")
	OR ("scenario_type" = "Data Entry" AND "Submitted Form" = "Yes")
        OR ("scenario_type" = "Data Entry" AND "Clicked Link?" = "Yes" AND "Submitted Form" = "No")
        OR ("scenario_type" = "Attachment" AND "Viewed Education?" = "Yes"))
	GROUP BY scenario_id
	ORDER BY component, date
'''



# Methods

# Workbook saving
def save():
        wb.save(filename = dest_xlsx)


def writeToSheet(header, sheet, results, summary=False):

	for col in range(0, len(header)):
		sheet.cell(row=1, column=col + 1).value = header[col]
	if summary:
		for row in results:
			sheet.append(row)
			components.append(row[0])
	else:
		for row in results:
			row = list(row)
			fixed_date = row[3].split(" ")[0]
			row[3] = fixed_date
			sheet.append(row)
	save()


# This is our first sheet
def overallStats():

	print("[+] Writing data to '{}'...".format(dest_xlsx))
	wb.active.title = "Response Rate All Components"
	sheet = wb.active
	header = ['component', 'response_rate']
	# All the data for the first worksheet
	# Ouputs "component, response_rate" (response rate is decimal rounded to nearest 1000th)
	query = '''
	with total as
		(SELECT scenario_id as id, COUNT("Email") as total
		FROM Reports
		WHERE ("Last Email Status" NOT LIKE "%bounced%")
		GROUP BY id
		),
	collection as
		(SELECT scenario_id, component as comp, CAST(COUNT("Email") AS FLOAT) / total.total as response_rate
		FROM Reports, total
		WHERE (total.id = scenario_id AND "Last Email Status" NOT LIKE "%bounced%")
        	AND (("scenario_type" = "Click-only" AND "Clicked Link?" = "Yes")
		OR ("scenario_type" = "Data Entry" AND "Submitted Form" = "Yes")
        	OR ("scenario_type" = "Data Entry" AND "Clicked Link?" = "Yes" AND "Submitted Form" = "No")
        	OR ("scenario_type" = "Attachment" AND "Viewed Education?" = "Yes"))
		GROUP BY scenario_id
		)

	SELECT comp, ROUND(AVG(collection.response_rate), 3)
	FROM collection
	GROUP BY comp
	'''
	results = c.execute(query)
	writeToSheet(header, sheet, results, 1)
	'''
	for col in range(0, len(header)):
		sheet.cell(row=1, column=col + 1).value = header[col]
	for row in results:
		sheet.append(row)
		components.append(row[0])
	save()
	'''


def overallStatsPerComp(name):

	header = ['scenario_id', 'component', 'scenario_type', 'date', 'emails_delivered', 'responses', 'response_rate']
	title = name + ' Overall Trend'
	sheet = wb.create_sheet(title=title)
	query = '''
        with total as
                (SELECT scenario_id as id, COUNT("Email") as total
                FROM Reports
                WHERE ("Last Email Status" NOT LIKE "%bounced%")
                GROUP BY id
                )

	SELECT scenario_id, component, scenario_type, "Last Email Status Timestamp" as date, total.total as total, COUNT("Email") as response, ROUND(CAST(COUNT("Email") AS FLOAT) / total.total, 3) as response_rate
        FROM Reports, total
        WHERE (component = '{}' AND total.id = scenario_id AND "Last Email Status" NOT LIKE "%bounced%")
        AND (("scenario_type" = "Click-only" AND "Clicked Link?" = "Yes")
        OR ("scenario_type" = "Data Entry" AND "Submitted Form" = "Yes")
        OR ("scenario_type" = "Data Entry" AND "Clicked Link?" = "Yes" AND "Submitted Form" = "No")
        OR ("scenario_type" = "Attachment" AND "Viewed Education?" = "Yes"))
        GROUP BY scenario_id
        ORDER BY component, date
        '''.format(name)
	results = c.execute(query)
	writeToSheet(header, sheet, results)


def clickOnlyStats(name):

	header = ['scenario_id', 'component', 'scenario_type', 'date', 'emails_delivered', 'responses', 'response_rate']
	title = name + ' Click-Only'
	sheet = wb.create_sheet(title=title)
	query = '''
	 with total as
                (SELECT scenario_id as id, COUNT("Email") as total
                FROM Reports
                WHERE ("Last Email Status" NOT LIKE "%bounced%")
                GROUP BY id
                )

	SELECT scenario_id, component, scenario_type, "Last Email Status Timestamp" as date, total.total as total, COUNT("Email") as response, ROUND(CAST(COUNT("Email") AS FLOAT) / total.total, 3) as response_rate
        FROM Reports, total
        WHERE (total.id = scenario_id AND "Last Email Status" NOT LIKE "%bounced%")
        AND ("component" = '{}' AND "scenario_type" = "Click-only" AND "Clicked Link?" = "Yes")
        GROUP BY scenario_id
	ORDER BY scenario_id
	'''.format(name)
	results = c.execute(query)
	writeToSheet(header, sheet, results)


def dataEntryStats(name):

	header = ['scenario_id', 'component', 'scenario_type', 'date', 'emails_delivered', 'clicked', 'submitted_form', 'response_rate']
	title = name + ' Data Entry'
	sheet = wb.create_sheet(title=title)
	query = '''
	with total as
                (SELECT scenario_id as id, COUNT("Email") as total
                FROM Reports
                WHERE ("Last Email Status" NOT LIKE "%bounced%")
                GROUP BY id
                ),
	clickedonly as
		(SELECT scenario_id as id, COUNT("Email") as total
		FROM Reports
		WHERE ("Last Email Status" NOT LIKE "%bounced%")
		AND ("scenario_type" = "Data Entry" AND "Clicked Link?" = "Yes" AND "Submitted Form" = "No")
		GROUP BY id
		),
	submitted as
                (SELECT scenario_id as id, COUNT("Email") as total
                FROM Reports
                WHERE ("Last Email Status" NOT LIKE "%bounced%")
		AND ("scenario_type" = "Data Entry" AND "Submitted Form" = "Yes")
                GROUP BY id
		)

	SELECT scenario_id, component, scenario_type, "Last Email Status Timestamp" as date, total.total as total, clickedonly.total as clicked, submitted.total as submitted_total, ROUND(CAST((clickedonly.total + submitted.total) AS FLOAT) / total.total, 3) as response_rate
        FROM Reports, total, clickedonly, submitted
        WHERE (total.id = scenario_id AND clickedonly.id = scenario_id AND submitted.id = scenario_id AND "Last Email Status" NOT LIKE "%bounced%")
        AND (component = '{}')
        GROUP BY scenario_id
	ORDER BY scenario_id
	'''.format(name)

	results = c.execute(query)
	writeToSheet(header, sheet, results)


def attachmentStats(name):

	header = ['scenario_id', 'component', 'scenario_type', 'date', 'emails_delivered', 'responses', 'response_rate']
	title = name + ' Attachment'
	sheet = wb.create_sheet(title=title)
	query = '''
        with total as
                (SELECT scenario_id as id, COUNT("Email") as total
                FROM Reports
                WHERE ("Last Email Status" NOT LIKE "%bounced%")
                GROUP BY id
                )

        SELECT scenario_id, component, scenario_type, "Last Email Status Timestamp" as date, total.total as total, COUNT("Email") as response, ROUND(CAST(COUNT("Email") AS FLOAT) / total.total, 3) as response_$
        FROM Reports, total
        WHERE (total.id = scenario_id AND "Last Email Status" NOT LIKE "%bounced%")
        AND (component = '{}' AND "scenario_type" = "Attachment" AND "Viewed Education?" = "Yes")
        GROUP BY scenario_id
        ORDER BY component, date
	'''.format(name)

	results = c.execute(query)
	writeToSheet(header, sheet, results)


def db_export():
	print('[+] Exporting database to csv...')
	data = c.execute('SELECT * FROM Reports')
	fieldnames = [
                # new fields that didn't exist
                'scenario_id', 'scenario_type', 'component',
                # shared by all
                'Email', 'Recipient Name', 'Recipient Group', 'Department', 'Location', 'Opened Email?', 'Opened Email Timestamp',
                # specific to attachement
                'Viewed Education?', 'Viewed Education Timestamp',
                # shared by clicked link, data entry
                'Clicked Link?', 'Clicked Link Timestamp',
                # specific to data entry
                'Submitted Form', 'Username', 'Entered Password?', 'Submitted Form Timestamp',
                # shared by all
                'Reported Phish?', 'New/Repeat Reporter', 'Reported Phish Timestamp', 'Time to Report (in seconds)', 'Remote IP',
                'GeoIP Country', 'GeoIP City', 'GeoIP Organization', 'Last DSN', 'Last Email Status',
                'Last Email Status Timestamp', 'Language', 'Browser', 'User-Agent', 'Mobile?',
                # specific to clicked link
                'Seconds Spent on Education Page', 'USDOJACCOUNTTYPE', 'SN', 'GIVENNAME', 'INITIALS',
                # specific to data entry
                'DUTY_STATION', 'RAND', 'Submitted Data']

	with open(export, 'w') as f:
		writer = csv.writer(f)
		writer.writerow(fieldnames)
		writer.writerows(data)

	print('[+] DONE. Database exported')
################
if __name__ == '__main__':


	if sys.argv[2] == '-e':
		db_export()
	elif sys.argv[2] == '-s':
		overallStats()

		for name in components:
			overallStatsPerComp(name)
			clickOnlyStats(name)
			dataEntryStats(name)
			attachmentStats(name)

		conn.close()
		print('[+] DONE. Spreadsheet Created.')
	else:
		print('''
			[!] USAGE: python3 report.py <db_name> <option>
				-s : create xlsx spreadsheet
				-e : export db to csv file
			''')
